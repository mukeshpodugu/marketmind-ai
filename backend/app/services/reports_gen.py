import io
import os
import datetime
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.config import settings

def generate_pdf_report(stock_data: dict, predictions: dict, sentiment: dict, indicators: dict) -> io.BytesIO:
    """
    Generates a beautifully styled financial PDF report containing stock forecasts,
    technical indicators, sentiment analysis, and risk metrics, with developer credits.
    """
    buffer = io.BytesIO()
    
    # Page setup
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )
    
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CoverTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=26,
        leading=32,
        textColor=colors.HexColor('#312E81'), # Deep Indigo
        spaceAfter=15
    )
    
    subtitle_style = ParagraphStyle(
        'CoverSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=12,
        leading=16,
        textColor=colors.HexColor('#4B5563'),
        spaceAfter=30
    )
    
    h1_style = ParagraphStyle(
        'SectionH1',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=16,
        leading=20,
        textColor=colors.HexColor('#1E1B4B'),
        spaceBefore=15,
        spaceAfter=10,
        keepWithNext=True
    )
    
    body_style = ParagraphStyle(
        'ReportBody',
        parent=styles['BodyText'],
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#1F2937'),
        spaceAfter=8
    )

    credit_style = ParagraphStyle(
        'CreditText',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#6B7280'),
        spaceAfter=3
    )

    story = []

    # --- Title Page / Header ---
    story.append(Paragraph("MarketMind AI – Financial Intelligence Report", title_style))
    story.append(Paragraph(f"Generated on {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", subtitle_style))
    
    # Developer Credit Card
    credit_data = [
        [Paragraph(f"<b>Analyst:</b> {settings.DEV_NAME}", body_style), Paragraph(f"<b>Location:</b> {settings.DEV_LOCATION}", body_style)],
        [Paragraph(f"<b>Email:</b> {settings.DEV_EMAIL}", body_style), Paragraph(f"<b>Phone:</b> {settings.DEV_PHONE}", body_style)],
        [Paragraph("<b>Platform:</b> MarketMind AI Research Platform", body_style), Paragraph("<b>Status:</b> Production Ready", body_style)]
    ]
    credit_table = Table(credit_data, colWidths=[3.5*inch, 3.5*inch])
    credit_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F3F4F6')),
        ('PADDING', (0,0), (-1,-1), 10),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LINEBELOW', (0,-1), (-1,-1), 1, colors.HexColor('#E5E7EB')),
        ('LINEABOVE', (0,0), (-1,0), 1, colors.HexColor('#E5E7EB')),
        ('LINELEFT', (0,0), (0,-1), 1, colors.HexColor('#E5E7EB')),
        ('LINERIGHT', (-1,0), (-1,-1), 1, colors.HexColor('#E5E7EB')),
    ]))
    story.append(credit_table)
    story.append(Spacer(1, 20))

    # --- Stock Summary ---
    story.append(Paragraph("1. Market Asset Overview", h1_style))
    symbol = stock_data.get("symbol", "N/A")
    name = stock_data.get("name", "N/A")
    sector = stock_data.get("sector", "N/A")
    curr_price = stock_data.get("current_price", 0.0)
    
    summary_text = (
        f"This investment brief reviews the historical trends and predictive forecast for <b>{name} ({symbol})</b>, "
        f"operating in the <b>{sector}</b> sector. The asset's current trading close price is recorded at "
        f"<b>${curr_price:.2f}</b>. Dynamic predictions were compiled using the selected optimal machine learning model."
    )
    story.append(Paragraph(summary_text, body_style))
    story.append(Spacer(1, 10))

    # --- Predictions ---
    story.append(Paragraph("2. Machine Learning Price Projections", h1_style))
    best_model = predictions.get("best_model", "LSTM")
    confidence = predictions.get("confidence_score", 0.85) * 100.0
    risk = predictions.get("risk_score", 5.0)
    risk_cat = predictions.get("risk_category", "Medium")
    
    model_text = (
        f"Our AI core evaluated six mathematical models: Linear Regression, Random Forest, XGBoost, and deep recurrent neural networks "
        f"(LSTM, GRU, Bi-LSTM). The champion model was identified as <b>{best_model}</b> with a model confidence score of "
        f"<b>{confidence:.1f}%</b>. Risk assessment indicates a <b>{risk_cat} Risk</b> profile (Score: {risk:.1f}/10)."
    )
    story.append(Paragraph(model_text, body_style))
    
    # Predictions Table
    pred_headers = [Paragraph("<b>Forecast Period</b>", body_style), Paragraph("<b>Projected Close Price</b>", body_style), Paragraph("<b>Confidence Range</b>", body_style), Paragraph("<b>Expected Trend</b>", body_style)]
    pred_rows = [pred_headers]
    
    for horizon, label in [("day", "Next Business Day"), ("week", "Next Week (t+5)"), ("month", "Next Month (t+22)"), ("quarter", "Next Quarter (t+66)")]:
        h_data = predictions.get("predictions", {}).get(horizon, [])
        if h_data:
            target = h_data[-1]
            price = target.get("predicted_price", 0.0)
            low = target.get("confidence_lower", 0.0)
            high = target.get("confidence_upper", 0.0)
            direction = target.get("direction", "up").upper()
            
            trend_color = "#10B981" if direction == "UP" else "#EF4444"
            trend_para = Paragraph(f"<font color='{trend_color}'><b>{direction}</b></font>", body_style)
            
            pred_rows.append([
                Paragraph(label, body_style),
                Paragraph(f"${price:.2f}", body_style),
                Paragraph(f"${low:.2f} - ${high:.2f}", body_style),
                trend_para
            ])
            
    pred_table = Table(pred_rows, colWidths=[2.2*inch, 1.8*inch, 1.8*inch, 1.2*inch])
    pred_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F9FAFB')]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E5E7EB')),
    ]))
    # Force textcolor for headers to white
    for i in range(len(pred_headers)):
        pred_headers[i].style.textColor = colors.white
        
    story.append(pred_table)
    story.append(Spacer(1, 15))

    # --- Technical Indicators ---
    story.append(Paragraph("3. Technical Indicator Signpost", h1_style))
    story.append(Paragraph("A breakdown of critical technical signals calculated over rolling periods:", body_style))
    
    ind_headers = [Paragraph("<b>Indicator</b>", body_style), Paragraph("<b>Current Metric</b>", body_style), Paragraph("<b>Signal Interpretation</b>", body_style)]
    ind_rows = [ind_headers]
    
    rsi = indicators.get("RSI", 50.0)
    rsi_sig = "Neutral"
    if rsi > 70: rsi_sig = "Overbought (Sell Signal)"
    elif rsi < 30: rsi_sig = "Oversold (Buy Signal)"
    
    macd = indicators.get("MACD", 0.0)
    macd_sig = "Bullish Crossover" if macd > indicators.get("MACD_Signal", 0.0) else "Bearish Crossover"
    
    ind_rows.append([Paragraph("RSI (14)", body_style), Paragraph(f"{rsi:.1f}", body_style), Paragraph(rsi_sig, body_style)])
    ind_rows.append([Paragraph("MACD Line", body_style), Paragraph(f"{macd:.4f}", body_style), Paragraph(macd_sig, body_style)])
    ind_rows.append([Paragraph("Volatility (Annualized)", body_style), Paragraph(f"{indicators.get('Volatility', 0.0)*100:.1f}%", body_style), Paragraph("Standard Risk Factor", body_style)])
    
    ind_table = Table(ind_rows, colWidths=[2.0*inch, 1.8*inch, 3.2*inch])
    ind_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#312E81')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F9FAFB')]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E5E7EB')),
    ]))
    for h in ind_headers:
        h.style.textColor = colors.white
        
    story.append(ind_table)
    story.append(Spacer(1, 15))

    # --- Sentiment ---
    story.append(Paragraph("4. News Intelligence & Sentiment Mood", h1_style))
    mood = sentiment.get("mood_index", 50.0)
    label = sentiment.get("sentiment_label", "Neutral")
    sent_text = (
        f"Processing of online financial news headlines yields a <b>Market Mood Index</b> of <b>{mood:.1f}/100</b>, "
        f"falling into the <b>{label}</b> category. This metric reflects short-term investor sentiment sentiment trends "
        f"mined via text classification algorithms."
    )
    story.append(Paragraph(sent_text, body_style))
    story.append(Spacer(1, 20))
    
    # Footer disclaimer page boundary
    story.append(Paragraph("<b>Disclaimer:</b> MarketMind AI is an educational predictive tool. Past performance does not guarantee future results. Consult a licensed financial advisor prior to making market trades.", credit_style))

    # Build PDF
    doc.build(story)
    buffer.seek(0)
    return buffer

def generate_excel_report(stock_data: dict, predictions: dict, sentiment: dict, indicators: dict) -> io.BytesIO:
    """
    Generates a beautifully styled financial Excel spreadsheet with sheets
    for Asset Overview, Forecasts, and Technical Signals.
    """
    wb = Workbook()
    
    # 1. Overview Sheet
    ws1 = wb.active
    ws1.title = "Asset Summary"
    ws1.views.sheetView[0].showGridLines = True
    
    # Styles
    font_title = Font(name="Calibri", size=16, bold=True, color="1E1B4B")
    font_section = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_normal = Font(name="Calibri", size=11)
    
    fill_indigo = PatternFill(start_color="312E81", end_color="312E81", fill_type="solid")
    fill_gray_light = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    fill_blue_header = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    
    thin_border_side = Side(border_style="thin", color="D1D5DB")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # Fill Title
    ws1["A1"] = "MarketMind AI – Analytical Profile"
    ws1["A1"].font = font_title
    
    # Developer credits
    ws1["A3"] = "Lead Developer:"
    ws1["A3"].font = font_bold
    ws1["B3"] = settings.DEV_NAME
    ws1["B3"].font = font_normal
    
    ws1["A4"] = "Email Contact:"
    ws1["A4"].font = font_bold
    ws1["B4"] = settings.DEV_EMAIL
    ws1["B4"].font = font_normal
    
    ws1["A5"] = "Location:"
    ws1["A5"].font = font_bold
    ws1["B5"] = settings.DEV_LOCATION
    ws1["B5"].font = font_normal
    
    # Stock info
    ws1["A7"] = "Stock Information"
    ws1.merge_cells("A7:B7")
    ws1["A7"].fill = fill_indigo
    ws1["A7"].font = font_section
    ws1["A7"].alignment = Alignment(horizontal="center")
    
    info_rows = [
        ("Company Name", stock_data.get("name")),
        ("Ticker Symbol", stock_data.get("symbol")),
        ("Sector", stock_data.get("sector")),
        ("Current Price ($)", stock_data.get("current_price")),
        ("Report Date", datetime.datetime.now().strftime("%Y-%m-%d")),
    ]
    
    row_idx = 8
    for label, val in info_rows:
        ws1.cell(row=row_idx, column=1, value=label).font = font_bold
        ws1.cell(row=row_idx, column=1).border = thin_border
        
        c = ws1.cell(row=row_idx, column=2, value=val)
        c.font = font_normal
        c.border = thin_border
        row_idx += 1

    # 2. Predictions Sheet
    ws2 = wb.create_sheet(title="AI Forecasts")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2["A1"] = f"Price Projections for {stock_data.get('symbol')}"
    ws2["A1"].font = font_title
    
    ws2["A3"] = "Champion Model:"
    ws2["A3"].font = font_bold
    ws2["B3"] = predictions.get("best_model")
    ws2["B3"].font = font_normal
    
    ws2["A4"] = "Model Confidence:"
    ws2["A4"].font = font_bold
    ws2["B4"] = predictions.get("confidence_score")
    ws2["B4"].font = font_normal
    ws2["B4"].number_format = '0.0%'
    
    # Table headers
    headers = ["Forecast Window", "Projected Price", "Confidence Lower Limit", "Confidence Upper Limit", "Directional Bias"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=6, column=col, value=h)
        cell.font = font_header
        cell.fill = fill_blue_header
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
    p_idx = 7
    for horizon, label in [("day", "Next Day"), ("week", "Next Week"), ("month", "Next Month"), ("quarter", "Next Quarter")]:
        h_data = predictions.get("predictions", {}).get(horizon, [])
        if h_data:
            target = h_data[-1]
            ws2.cell(row=p_idx, column=1, value=label).font = font_bold
            ws2.cell(row=p_idx, column=2, value=target.get("predicted_price")).number_format = '$#,##0.00'
            ws2.cell(row=p_idx, column=3, value=target.get("confidence_lower")).number_format = '$#,##0.00'
            ws2.cell(row=p_idx, column=4, value=target.get("confidence_upper")).number_format = '$#,##0.00'
            ws2.cell(row=p_idx, column=5, value=target.get("direction").upper()).font = font_normal
            
            for col in range(1, 6):
                ws2.cell(row=p_idx, column=col).border = thin_border
                
            p_idx += 1

    # Auto-fit columns
    for ws in [ws1, ws2]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
